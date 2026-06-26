"""
shared_utils/riddles.py
Robust reading of the "Cultural Riddles Benchmark" .xlsx files.

The source files are not uniform:
  - the riddle sheet is usually named "Riddles" but the 4 Uganda files name their
    single sheet "Cultural Riddles Benchmark Luga/luma/Luso/Runy";
  - column ORDER varies (e.g. Bengali_India interleaves the answer columns).

So we resolve the sheet by name (fallback: the sole sheet) and resolve columns by
header TEXT rather than fixed position.
"""

import re
from typing import Dict, List, Optional

import openpyxl


class RiddleSheetError(Exception):
    """Raised when a workbook has no resolvable riddle sheet / required columns."""


# Required logical fields -> predicate over the (stripped) header cell text.
_FIELD_MATCHERS = {
    "topic":               lambda h: h == "topic",
    "riddle_original":     lambda h: h.startswith("riddle (original"),
    "riddle_translation":  lambda h: h.startswith("riddle (english"),
    "ref_orig":            lambda h: h.startswith("reference answer (original"),
    "ref_en":              lambda h: h.startswith("reference answer (english"),
}
# Optional fields (nice to keep, not required for validation).
_OPTIONAL_MATCHERS = {
    "number": lambda h: h == "number",
    "author": lambda h: h == "author",
}

REQUIRED_FIELDS = list(_FIELD_MATCHERS.keys())


def resolve_riddle_sheet(wb) -> str:
    """
    Pick the worksheet holding the riddles.
      1. a sheet whose name contains "riddle" (case-insensitive), else
      2. the sole sheet if the workbook has exactly one, else
      3. raise RiddleSheetError.
    """
    named = [n for n in wb.sheetnames if "riddle" in n.lower()]
    if named:
        return named[0]
    if len(wb.sheetnames) == 1:
        return wb.sheetnames[0]
    raise RiddleSheetError(
        f"no sheet name contains 'riddle' and workbook has "
        f"{len(wb.sheetnames)} sheets: {wb.sheetnames}"
    )


def _norm(cell) -> str:
    return str(cell).strip().lower() if cell is not None else ""


def resolve_columns(header_row) -> Dict[str, int]:
    """
    Map logical field -> column index by matching header text.
    Raises RiddleSheetError if any REQUIRED field is missing.
    """
    cols: Dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        h = _norm(cell)
        if not h:
            continue
        for field, matches in {**_FIELD_MATCHERS, **_OPTIONAL_MATCHERS}.items():
            if field not in cols and matches(h):
                cols[field] = idx
    missing = [f for f in REQUIRED_FIELDS if f not in cols]
    if missing:
        raise RiddleSheetError(f"missing required column(s): {missing}")
    return cols


def _find_header_row(ws, max_scan: int = 8):
    """Return (header_values, data_start_row_index) by scanning the first rows."""
    rows = list(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True))
    for i, row in enumerate(rows):
        norm = {_norm(c) for c in row}
        if "topic" in norm and any(
            (str(c).strip().lower().startswith("riddle (")) for c in row if c is not None
        ):
            return row, i + 1  # header is 1-based row (i+1)
    # Fall back to first row.
    return (rows[0] if rows else []), 1


def read_riddles_xlsx(path: str) -> List[Dict[str, Optional[str]]]:
    """
    Read one riddles workbook and return a list of row dicts with keys:
      number, topic, author, riddle_original, riddle_translation, ref_orig, ref_en
    Only rows that have at least one of riddle_original / riddle_translation are kept.
    Raises RiddleSheetError on unresolvable sheet / missing columns.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = resolve_riddle_sheet(wb)
        ws = wb[sheet]
        header, data_start = _find_header_row(ws)
        cols = resolve_columns(header)

        def get(row, field):
            idx = cols.get(field)
            if idx is None or idx >= len(row):
                return None
            v = row[idx]
            if v is None:
                return None
            s = str(v).strip()
            return s or None

        out: List[Dict[str, Optional[str]]] = []
        for row in ws.iter_rows(min_row=data_start + 1, values_only=True):
            rec = {f: get(row, f) for f in
                   ("number", "topic", "author", "riddle_original",
                    "riddle_translation", "ref_orig", "ref_en")}
            if rec["riddle_original"] or rec["riddle_translation"]:
                out.append(rec)
        return out
    finally:
        wb.close()


_BRACKET_RE = re.compile(r"\[(.*)\]")


def parse_lang_region_key(filename: str) -> Optional[str]:
    """Extract the lang_region key from a 'Cultural Riddles Benchmark [KEY].xlsx' name."""
    m = _BRACKET_RE.search(filename)
    return m.group(1).strip() if m else None
