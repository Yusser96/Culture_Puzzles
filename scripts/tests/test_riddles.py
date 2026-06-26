"""Tests for shared_utils/riddles.py — sheet/column resolution and xlsx reading."""

import os
import tempfile
import unittest

from tests.helpers import HEADER, HEADER_REORDERED, make_workbook

from shared_utils.riddles import (
    resolve_riddle_sheet, resolve_columns, read_riddles_xlsx,
    parse_lang_region_key, RiddleSheetError,
)
import openpyxl


def _rows(header, body):
    return [header] + body


class TestParseKey(unittest.TestCase):
    def test_basic(self):
        self.assertEqual(
            parse_lang_region_key("Cultural Riddles Benchmark [Arabic_Egypt].xlsx"),
            "Arabic_Egypt",
        )

    def test_duplicate_suffix_collides(self):
        # The "(1)" is outside the brackets, so both files yield the same key.
        self.assertEqual(
            parse_lang_region_key("Cultural Riddles Benchmark [Spanish_Argentina](1).xlsx"),
            "Spanish_Argentina",
        )

    def test_space_and_emoji_preserved(self):
        self.assertEqual(
            parse_lang_region_key("Cultural Riddles Benchmark [Sinhala_Sri Lanka].xlsx"),
            "Sinhala_Sri Lanka",
        )

    def test_no_brackets(self):
        self.assertIsNone(parse_lang_region_key("no_brackets_here.xlsx"))


class TestSheetResolution(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.mkdtemp()

    def _path(self, name="wb.xlsx"):
        return os.path.join(self.tmp, name)

    def test_named_riddles_sheet(self):
        p = self._path()
        make_workbook(p, [("Instructions", [["x"]]), ("Riddles", _rows(HEADER, []))])
        wb = openpyxl.load_workbook(p, read_only=True)
        self.assertEqual(resolve_riddle_sheet(wb), "Riddles")

    def test_single_sheet_fallback(self):
        # Uganda-style: one oddly named sheet, no "Riddles".
        p = self._path()
        make_workbook(p, [("Cultural Riddles Benchmark Luga", _rows(HEADER, []))])
        wb = openpyxl.load_workbook(p, read_only=True)
        # name contains "riddle" -> picked by rule 1
        self.assertEqual(resolve_riddle_sheet(wb), "Cultural Riddles Benchmark Luga")

    def test_single_sheet_no_riddle_in_name(self):
        p = self._path()
        make_workbook(p, [("Sheet1", _rows(HEADER, []))])
        wb = openpyxl.load_workbook(p, read_only=True)
        self.assertEqual(resolve_riddle_sheet(wb), "Sheet1")

    def test_unresolvable_multisheet(self):
        p = self._path()
        make_workbook(p, [("Instructions", [["x"]]), ("Data", [["y"]])])
        wb = openpyxl.load_workbook(p, read_only=True)
        with self.assertRaises(RiddleSheetError):
            resolve_riddle_sheet(wb)


class TestColumnResolution(unittest.TestCase):
    def test_canonical_order(self):
        cols = resolve_columns(HEADER)
        self.assertEqual(cols["riddle_original"], 3)
        self.assertEqual(cols["riddle_translation"], 4)
        self.assertEqual(cols["ref_orig"], 7)
        self.assertEqual(cols["ref_en"], 8)

    def test_reordered(self):
        cols = resolve_columns(HEADER_REORDERED)
        self.assertEqual(cols["riddle_original"], 3)
        self.assertEqual(cols["ref_orig"], 4)
        self.assertEqual(cols["riddle_translation"], 5)
        self.assertEqual(cols["ref_en"], 6)

    def test_missing_required(self):
        with self.assertRaises(RiddleSheetError):
            resolve_columns(["Number", "Topic", "Author"])


class TestReadRiddles(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.mkdtemp()

    def test_reads_and_filters_empty_rows(self):
        p = os.path.join(self.tmp, "wb.xlsx")
        body = [
            [1, "Politics", "A", "orig1", "trans1", "", "", "ans1", "ansE1"],
            [2, "Arts", "B", "orig2", "trans2", "", "", "ans2", "ansE2"],
            [3, "Arts", "B", None, None, "", "", None, None],   # empty -> skipped
        ]
        make_workbook(p, [("Instructions", [["x"]]), ("Riddles", _rows(HEADER, body))])
        rows = read_riddles_xlsx(p)
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0]["riddle_original"], "orig1")
        self.assertEqual(rows[0]["ref_en"], "ansE1")
        self.assertEqual(rows[1]["topic"], "Arts")

    def test_reordered_columns_read_correctly(self):
        p = os.path.join(self.tmp, "wb.xlsx")
        body = [[1, "Politics", "A", "ORIG", "REFO", "TRANS", "REFE"]]
        make_workbook(p, [("Riddles", _rows(HEADER_REORDERED, body))])
        rows = read_riddles_xlsx(p)
        self.assertEqual(rows[0]["riddle_original"], "ORIG")
        self.assertEqual(rows[0]["riddle_translation"], "TRANS")
        self.assertEqual(rows[0]["ref_orig"], "REFO")
        self.assertEqual(rows[0]["ref_en"], "REFE")

    def test_missing_columns_raises(self):
        p = os.path.join(self.tmp, "wb.xlsx")
        make_workbook(p, [("Riddles", _rows(["Number", "Topic"], [[1, "Politics"]]))])
        with self.assertRaises(RiddleSheetError):
            read_riddles_xlsx(p)


if __name__ == "__main__":
    unittest.main()
